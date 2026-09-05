#!/usr/bin/env python3
"""
Extractor TONTO para el seed de asesorías. Sin lógica de negocio a propósito:
solo saca los datos del Excel y del PDF de la facultad a archivos que el seed
(`src/db/seed/asesorias.ts`) sabe leer. Todo lo semántico —empatar docentes,
resolver cursos, detectar desfases— vive en TypeScript, donde está probado.

    python3 scripts/asesorias-extraer.py <atencion.xlsx> <horario.pdf> [--periodo 2026-2]

Escribe en scripts/out/ (gitignored: el Excel trae enlaces Zoom personales de
más de cien profesores y el repo es público):
    scripts/out/asesorias-<periodo>.json   filas del Excel, celdas tal cual
    scripts/out/horario-<periodo>.txt      texto plano del PDF

Necesita openpyxl y pypdf (ambos instalados en esta Mac).
"""
import json, sys, os
from pathlib import Path

args = [a for a in sys.argv[1:] if not a.startswith("--")]
periodo = next((sys.argv[i + 1] for i, a in enumerate(sys.argv) if a == "--periodo"), "2026-2")
if len(args) != 2:
    print(__doc__); sys.exit(2)
xlsx, pdf = args
out = Path(__file__).resolve().parent / "out"
out.mkdir(exist_ok=True)

import openpyxl
from pypdf import PdfReader

wb = openpyxl.load_workbook(xlsx, data_only=True)

def hoja(nombre):
    for ws in wb.worksheets:
        if ws.title.strip().lower() == nombre:
            return ws
    raise SystemExit(f"No encuentro la hoja '{nombre}' en {xlsx}; hay: {wb.sheetnames}")

# Atención alumnos. Docente y Zoom vienen en celdas COMBINADAS por profesor
# (46 bloques de Zoom coinciden exactamente con bloques de docente): openpyxl
# solo devuelve el valor en la celda superior y None en el resto. Expandirlo a
# todo el rango combinado es representación del Excel, no lógica de negocio.
ws = hoja("atención alumnos")
expandido = {}   # (fila, col) -> valor de la celda superior del rango combinado
for rango in ws.merged_cells.ranges:
    valor = ws.cell(rango.min_row, rango.min_col).value
    for f in range(rango.min_row, rango.max_row + 1):
        expandido[(f, rango.min_col)] = valor
atencion, docente = [], None
for i, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
    doc, asig, dia, ini, fin, amb, zoom = (tuple(row) + (None,) * 7)[:7]
    doc = doc if doc is not None else expandido.get((i, 1))
    zoom = zoom if zoom is not None else expandido.get((i, 7))
    if isinstance(doc, str) and doc.strip():
        docente = doc.strip()
    if asig is None and dia is None and ini is None:
        continue
    if str(dia).strip().lower() == "día":     # fila de cabecera repetida
        continue
    atencion.append({"fila": i, "docente": docente, "asignatura": asig, "dia": dia,
                     "inicio": ini, "fin": fin, "ambiente": amb, "zoom": zoom})

tesis = []
for i, row in enumerate(hoja("asesoría de tesis").iter_rows(min_row=3, values_only=True), start=3):
    dia, doc, ini, fin, amb, zoom = (tuple(row) + (None,) * 6)[:6]
    if doc and str(dia).strip().lower() != "día":
        tesis.append({"fila": i, "docente": doc, "dia": dia, "inicio": ini, "fin": fin,
                      "ambiente": amb, "zoom": zoom})

json.dump({"periodo": periodo, "origen": os.path.basename(xlsx), "atencion": atencion, "tesis": tesis},
          open(out / f"asesorias-{periodo}.json", "w"), ensure_ascii=False, indent=1, default=str)

texto = "\n".join((p.extract_text() or "") for p in PdfReader(pdf).pages)
(out / f"horario-{periodo}.txt").write_text(texto, encoding="utf-8")

print(f"atención: {len(atencion)} filas · tesis: {len(tesis)} filas · PDF: {texto.count(chr(10))} líneas")
print(f"-> {out}/asesorias-{periodo}.json  y  {out}/horario-{periodo}.txt")
